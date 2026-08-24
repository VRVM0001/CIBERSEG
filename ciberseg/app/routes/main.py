"""
=========================================================================
 Módulo: main.py (app/routes/main.py)
 Rutas principales: dashboard con métricas reales y placeholders de módulos.

 Este blueprint agrupa todo lo que no pertenece a un módulo de negocio
 concreto: el dashboard (la pantalla más consultada del sistema), el
 buscador global, las notificaciones de la campana, la ayuda, el perfil
 del usuario, la configuración del sistema y las exportaciones del
 dashboard a Excel y PDF.
=========================================================================
"""
import datetime      # Para calcular fechas: hoy, y los ultimos N meses de las graficas
from flask import Blueprint, render_template, abort

from ..db import query_all, query_one, db_status   # Acceso a datos definido en db.py

bp = Blueprint("main", __name__)
# Crea el blueprint "main". Sus rutas se invocan despues como
# url_for("main.index"), url_for("main.configuracion"), etc.

# Meses abreviados en español
# Se usa para traducir el numero de mes (1-12) a su abreviatura, que es
# lo que se muestra en el eje horizontal de las graficas del dashboard.
MES_ABBR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


"""
   Función: ultimos_meses
   Objetivo: Calcular la lista de los últimos N meses (año y mes),
             contando hacia atrás desde el mes actual, y devolverla en
             orden cronológico (el más antiguo primero).
   Parámetros:
     - n: cuántos meses hacia atrás calcular (por defecto 6).
   Retorno: Lista de tuplas (año, mes). Por ejemplo, si hoy es julio
            de 2026 y n=6: [(2026,2), (2026,3), ..., (2026,7)].
"""
def ultimos_meses(n=6):
    """Devuelve [(año, mes), ...] de los últimos n meses, en orden cronológico."""
    hoy = datetime.date.today()
    y, m = hoy.year, hoy.month          # Se parte del año y mes actuales
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1                          # Retrocede un mes
        if m == 0:                      # Si se paso de enero (el mes 0 no existe)...
            m, y = 12, y - 1            # ...vuelve a diciembre del año anterior
    out.reverse()                       # Se calculo del mas reciente al mas antiguo: se invierte
    return out


"""
   Función: index
   Objetivo: Ser la ruta del dashboard (la pantalla de inicio del
             sistema). Calcula los cuatro indicadores principales
             (KPIs), las series de datos de las cuatro gráficas y los
             dos rankings, y se los entrega a la plantilla.
   Parámetros: No recibe.
   Retorno: Si la base de datos no responde, la plantilla de "sin
            conexión". Si responde, la plantilla dashboard.html con
            todos los datos ya calculados.
"""
@bp.route("/")
def index():
    estado = db_status()
    if not estado["conectada"]:
        return render_template("sin_conexion.html", estado=estado, active="dashboard")
    # Si la base no responde se corta aqui: todas las consultas de abajo
    # fallarian igual, y asi el usuario ve un mensaje claro en vez de un error.

    # ---------- Filtros de la barra superior ----------
    # Los dos desplegables del dashboard envian su valor por GET y recargan
    # la pagina. Aqui se leen, se validan y se devuelven a la plantilla para
    # que el desplegable quede marcado en la opcion elegida.
    try:
        meses_sel = int(request.args.get("meses") or 6)
    except ValueError:
        meses_sel = 6                       # Si llegara texto en vez de numero
    if meses_sel not in (3, 6, 12):
        meses_sel = 6                       # Solo se aceptan los tres valores del desplegable
    # Validar contra una lista cerrada evita que un valor manipulado en la
    # URL (?meses=9999) genere consultas absurdas.

    try:
        cliente_sel = int(request.args.get("cliente") or 0) or None
    except ValueError:
        cliente_sel = None                  # None = "Todos los clientes"

    clientes_lista = query_all(
        "SELECT c.id, e.nombre FROM clientes c JOIN empresas e ON e.id = c.empresa_id "
        "ORDER BY e.nombre")
    # Alimenta el desplegable de clientes. El nombre visible esta en
    # empresas, por eso el JOIN.

    if cliente_sel and not any(c["id"] == cliente_sel for c in clientes_lista):
        cliente_sel = None                  # Id inexistente: se ignora el filtro

    f_cli = "AND cliente_id = ? " if cliente_sel else ""
    p_cli = [cliente_sel] if cliente_sel else []
    # Fragmento de SQL y su parametro, que se van agregando a las consultas
    # que dependen del cliente. Si no hay filtro, ambos quedan vacios y las
    # consultas se comportan igual que antes.

    # ---------- KPIs ----------
    clientes_activos = query_one("SELECT COUNT(*) n FROM clientes WHERE estado='Activo'")["n"]
    clientes_total   = query_one("SELECT COUNT(*) n FROM clientes")["n"]
    # Estos dos NO se filtran por cliente: representan el tamano de la
    # cartera completa, que no cambia al mirar un cliente concreto.
    equipos_oper     = query_one(
        "SELECT COUNT(*) n FROM equipos WHERE estado='Operativo' " + f_cli, p_cli)["n"]
    equipos_total    = query_one("SELECT COUNT(*) n FROM equipos WHERE 1=1 " + f_cli, p_cli)["n"]
    cotiz_pend       = query_one(
        "SELECT COUNT(*) n FROM cotizaciones WHERE estado IN ('Borrador','Enviada') "
        + f_cli, p_cli)["n"]
    cotiz_total      = query_one(
        "SELECT COUNT(*) n FROM cotizaciones WHERE 1=1 " + f_cli, p_cli)["n"]
    # "WHERE 1=1" permite concatenar el AND del filtro sin comprobar si es
    # la primera condicion. Es el mismo patron que usan los listados.
    # Seis consultas de conteo, una por cada numero de las tarjetas del
    # dashboard. Cada una usa COUNT(*), que el motor resuelve muy rapido.

    # Ingresos del periodo seleccionado (facturas pagadas)
    ing_mes = query_one(
        "SELECT COALESCE(SUM(total),0) s FROM facturas "
        "WHERE estado='Pagada' "
        "AND fecha_emision >= DATEADD(MONTH,?,CAST(GETDATE() AS DATE)) " + f_cli,
        [-meses_sel] + p_cli
    )["s"]
    # Ingresos del periodo anterior (para calcular la variación %)
    ing_mes_ant = query_one(
        "SELECT COALESCE(SUM(total),0) s FROM facturas "
        "WHERE estado='Pagada' "
        "AND fecha_emision >= DATEADD(MONTH,?,CAST(GETDATE() AS DATE)) "
        "AND fecha_emision <  DATEADD(MONTH,?,CAST(GETDATE() AS DATE)) " + f_cli,
        [-meses_sel * 2, -meses_sel] + p_cli
    )["s"]
    # Las dos consultas anteriores traen dos ventanas del mismo tamano: la
    # actual (ultimos N meses) y la inmediatamente anterior (de 2N a N
    # meses atras), para poder comparar una contra otra.
    # El numero de meses viaja como PARAMETRO, no concatenado en el texto:
    # DATEADD(MONTH,?,...) recibe el valor ya en negativo desde la lista.
    delta_ing = None
    if ing_mes_ant and ing_mes_ant > 0:      # Solo se calcula si hubo ingresos antes (evita dividir por cero)
        delta_ing = round((float(ing_mes) - float(ing_mes_ant)) / float(ing_mes_ant) * 100)
        # Formula de variacion porcentual: (actual - anterior) / anterior * 100

    kpis = {
        "clientes_activos": clientes_activos,
        "clientes_total": clientes_total,
        "ingresos_mes": float(ing_mes),
        "delta_ing": delta_ing,
        "cotiz_pend": cotiz_pend,
        "cotiz_total": cotiz_total,
        "equipos_oper": equipos_oper,
        "equipos_total": equipos_total,
    }

    # ---------- Series por mes (periodo seleccionado) ----------
    meses = ultimos_meses(meses_sel)                          # [(año,mes), ...] en orden cronologico
    labels = [MES_ABBR[m - 1] for (y, m) in meses]            # ["Feb","Mar",...] para el eje de la grafica
    claves = [f"{y:04d}-{m:02d}" for (y, m) in meses]         # ["2026-02","2026-03",...] para cruzar con la BD

    fv = {r["m"]: float(r["t"]) for r in query_all(
        "SELECT FORMAT(fecha_emision,'yyyy-MM') m, SUM(total) t FROM facturas "
        "WHERE fecha_emision >= DATEADD(MONTH,?,GETDATE()) " + f_cli +
        "GROUP BY FORMAT(fecha_emision,'yyyy-MM')", [-meses_sel] + p_cli)}
    # FORMAT(fecha,'yyyy-MM') convierte la fecha en texto "2026-06" para
    # poder agrupar por mes. El resultado queda en un diccionario fv
    # con la forma {"2026-06": 15000.0, ...}.
    ventas_mes = [round(fv.get(k, 0)) for k in claves]
    # Se recorren TODAS las claves esperadas y se rellena con 0 los meses
    # sin ventas: asi la grafica no queda con huecos ni se desalinea
    # respecto a las etiquetas.

    cv = {r["m"]: r["c"] for r in query_all(
        "SELECT FORMAT(fecha,'yyyy-MM') m, COUNT(*) c FROM cotizaciones "
        "WHERE fecha >= DATEADD(MONTH,?,GETDATE()) " + f_cli +
        "GROUP BY FORMAT(fecha,'yyyy-MM')", [-meses_sel] + p_cli)}
    cotiz_mes = [cv.get(k, 0) for k in claves]     # Mismo patron de relleno con ceros

    # ---------- Donuts ----------
    colores_eq = {"Firewall": "#dc2626", "Switch": "#2563eb",
                  "Access Point": "#16a34a", "Servidor": "#9333ea", "Otro": "#6b7280"}
    # Un color fijo por tipo de equipo, para que el donut se vea siempre
    # igual sin importar el orden en que vengan los datos.
    equipos_tipo = [
        dict(r, color=colores_eq.get(r["nombre"], "#6b7280"))
        for r in query_all(
            "SELECT tipo nombre, COUNT(*) n FROM equipos WHERE 1=1 " + f_cli +
            "GROUP BY tipo ORDER BY n DESC", p_cli)
    ]
    # dict(r, color=...) copia cada fila del resultado y le agrega la
    # clave "color" segun el tipo, que es lo que la grafica necesita.
    clientes_tipo = query_all(
        "SELECT tipo nombre, COUNT(*) n FROM clientes GROUP BY tipo ORDER BY n DESC")

    # ---------- Top listas ----------
    top_productos = query_all(
        "SELECT TOP 5 pr.nombre, SUM(d.cantidad) v FROM detalle_cotizacion d "
        "JOIN productos pr ON pr.id=d.producto_id "
        "JOIN cotizaciones co ON co.id=d.cotizacion_id WHERE 1=1 "
        + ("AND co.cliente_id = ? " if cliente_sel else "") +
        "GROUP BY pr.nombre ORDER BY v DESC", p_cli)
    # Se agrega el JOIN a cotizaciones porque detalle_cotizacion no guarda
    # el cliente: hay que subir a la cabecera para saber de quien es la linea.
    top_clientes = query_all(
        "SELECT TOP 5 e.nombre, SUM(f.total) v FROM facturas f "
        "JOIN clientes c ON c.id=f.cliente_id JOIN empresas e ON e.id=c.empresa_id "
        "WHERE 1=1 " + ("AND f.cliente_id = ? " if cliente_sel else "") +
        "GROUP BY e.nombre ORDER BY v DESC", p_cli)
    # Aqui el filtro lleva el alias f. delante, porque en esta consulta hay
    # tres tablas y 'cliente_id' sin alias seria ambiguo.
    # Doble JOIN encadenado (facturas -> clientes -> empresas) porque el
    # nombre visible del cliente esta en la tabla empresas, no en clientes.

    return render_template(
        "dashboard.html",
        active="dashboard",
        kpis=kpis,
        chart_labels=labels,
        ventas_mes=ventas_mes,
        cotiz_mes=cotiz_mes,
        equipos_tipo=equipos_tipo,
        clientes_tipo=[dict(r) for r in clientes_tipo],
        top_productos=[dict(r, v=int(r["v"])) for r in top_productos],
        top_clientes=[dict(r, v=float(r["v"])) for r in top_clientes],
        clientes_lista=[dict(r) for r in clientes_lista],
        cliente_sel=cliente_sel,
        meses_sel=meses_sel,
        # Estas tres son las que consume la barra de filtros de
        # dashboard.html: la lista del desplegable y los dos valores
        # elegidos, para dejar marcada la opcion activa.
    )
    # dict(r, v=...) convierte los valores numericos (que pyodbc puede
    # entregar como Decimal) a int o float, para que Jinja2 los pueda
    # formatear sin problemas en la plantilla.


# Información de los módulos aún no implementados (placeholder)
# Diccionario vacio: quedo reservado para modulos futuros que aun no
# tienen su propio blueprint. Mientras este vacio, la ruta modulo() de
# abajo responde 404 para cualquier nombre.
MODULOS_INFO = {
    }


"""
   Función: modulo
   Objetivo: Mostrar una pantalla genérica de "módulo en construcción"
             para cualquier nombre registrado en MODULOS_INFO.
   Parámetros:
     - nombre: identificador del módulo, que llega en la URL
               (/modulo/<nombre>).
   Retorno: La plantilla placeholder.html si el nombre está
            registrado; un error 404 si no lo está.
"""
@bp.route("/modulo/<nombre>")
def modulo(nombre):
    info = MODULOS_INFO.get(nombre)
    if not info:
        abort(404)                # Nombre no registrado: error "pagina no encontrada"
    titulo, fase = info
    return render_template("placeholder.html", active=nombre, titulo=titulo, fase=fase)


# =====================================================================
#  BARRA SUPERIOR: búsqueda global, notificaciones y ayuda (Fase 3)
# =====================================================================
from flask import request, jsonify, session  # noqa: E402
from ..db import query_all, query_one  # noqa: E402,F811
# Estos imports estan aqui abajo y no al inicio del archivo porque esta
# seccion se agrego en una fase posterior del proyecto. Los comentarios
# noqa le indican al revisor de estilo que ignore el import fuera de
# orden (E402) y la redefinicion (F811): ambas son intencionales.


"""
   Función: buscar
   Objetivo: Implementar el buscador global de la barra superior:
             consulta simultáneamente en empresas, clientes, contactos,
             cotizaciones y facturas, y reúne los resultados de las
             cinco entidades en una sola pantalla.
   Parámetros: No recibe directamente (lee el texto 'q' de la URL).
   Retorno: La plantilla buscar.html con un diccionario de resultados
            agrupados por entidad y el total de coincidencias.
"""
@bp.route("/buscar")
def buscar():
    """Búsqueda global: empresas, clientes, contactos, cotizaciones y facturas."""
    q = (request.args.get("q") or "").strip()
    resultados = {}
    if len(q) >= 2:      # Evita buscar con una sola letra, que traeria demasiados resultados
        like = f"%{q}%"
        resultados["empresas"] = query_all(
            "SELECT TOP 8 id, nombre, sector, ciudad FROM empresas "
            "WHERE nombre LIKE ? OR rnc LIKE ? ORDER BY nombre", [like, like])
        resultados["clientes"] = query_all(
            "SELECT TOP 8 c.id, c.codigo, c.estado, e.nombre "
            "FROM clientes c JOIN empresas e ON e.id = c.empresa_id "
            "WHERE e.nombre LIKE ? OR c.codigo LIKE ? ORDER BY e.nombre", [like, like])
        resultados["contactos"] = query_all(
            "SELECT TOP 8 ct.id, ct.nombre, ct.cargo, e.nombre AS empresa "
            "FROM contactos ct JOIN empresas e ON e.id = ct.empresa_id "
            "WHERE ct.nombre LIKE ? OR ct.email LIKE ? ORDER BY ct.nombre", [like, like])
        resultados["cotizaciones"] = query_all(
            "SELECT TOP 8 co.id, co.numero, co.estado, co.total, e.nombre AS cliente "
            "FROM cotizaciones co JOIN clientes c ON c.id = co.cliente_id "
            "JOIN empresas e ON e.id = c.empresa_id "
            "WHERE co.numero LIKE ? OR e.nombre LIKE ? ORDER BY co.id DESC", [like, like])
        resultados["facturas"] = query_all(
            "SELECT TOP 8 f.id, f.numero, f.estado, f.total, e.nombre AS cliente "
            "FROM facturas f JOIN clientes c ON c.id = f.cliente_id "
            "JOIN empresas e ON e.id = c.empresa_id "
            "WHERE f.numero LIKE ? OR e.nombre LIKE ? ORDER BY f.id DESC", [like, like])
        # TOP 8 en cada una de las cinco consultas: limita cuantos
        # resultados se muestran por categoria para no saturar la pantalla.
    total = sum(len(v) for v in resultados.values())    # Suma las coincidencias de las cinco entidades
    return render_template("buscar.html", active="", q=q,
                           resultados=resultados, total=total)


"""
   Función: api_notificaciones
   Objetivo: Calcular en tiempo real los avisos de la campana de la
             barra superior: cotizaciones esperando respuesta,
             facturas vencidas, equipos con problemas, stock bajo,
             licencias por vencer, seguimientos pendientes y
             prospectos sin atender. Cada tipo de aviso solo se
             calcula si está activado en la configuración.
   Parámetros: No recibe (consulta la tabla configuracion internamente).
   Retorno: Una respuesta JSON con el total de avisos y la lista de
            cada uno, con su texto, la URL a la que lleva y su tipo
            ("info" o "alerta").
"""
@bp.route("/api/notificaciones")
def api_notificaciones():
    """Notificaciones reales calculadas desde la base de datos."""
    from ..configuracion import leer_config
    cfg = leer_config()          # Trae todos los parametros, incluidos los interruptores notif_*
    avisos = []
    if cfg["notif_activas"] != "1":
        return jsonify({"total": 0, "avisos": []})
        # Interruptor general apagado: se devuelve vacio sin gastar
        # ni una sola consulta a la base de datos.
    n = query_one("SELECT COUNT(*) n FROM cotizaciones WHERE estado = 'Enviada'")["n"] \
        if cfg["notif_cotizaciones"] == "1" else 0
    # Patron que se repite en los siete bloques siguientes: la consulta
    # SOLO se ejecuta si ese tipo de notificacion esta encendido; si
    # esta apagado, n vale 0 sin ir a la base.
    if n:
        avisos.append({"texto": f"{n} cotización(es) enviadas esperando respuesta",
                       "url": "/cotizaciones?estado=Enviada", "tipo": "info"})
    n = query_one(
        "SELECT COUNT(*) n FROM facturas WHERE estado = 'Pendiente' "
        "AND fecha_vencimiento < CAST(GETDATE() AS DATE)")["n"] \
        if cfg["notif_facturas"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} factura(s) pendientes ya vencidas",
                       "url": "/facturas?estado=Pendiente", "tipo": "alerta"})
    n = query_one("SELECT COUNT(*) n FROM facturas WHERE estado = 'Vencida'")["n"] \
        if cfg["notif_facturas"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} factura(s) en estado Vencida",
                       "url": "/facturas?estado=Vencida", "tipo": "alerta"})
    n = query_one(
        "SELECT COUNT(*) n FROM equipos WHERE estado IN "
        "('Fuera de servicio','En mantenimiento')")["n"] \
        if cfg["notif_equipos"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} equipo(s) fuera de servicio o en mantenimiento",
                       "url": "/equipos", "tipo": "alerta"})
    n = query_one(
        "SELECT COUNT(*) n FROM productos WHERE activo = 1 "
        "AND tipo = 'Hardware' AND stock <= stock_minimo")["n"] \
        if cfg["notif_inventario"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} producto(s) con stock en el mínimo o por debajo",
                       "url": "/productos", "tipo": "alerta"})
    n = query_one(
        "SELECT COUNT(*) n FROM licencias WHERE estado = 'Activa' "
        "AND fecha_fin IS NOT NULL "
        "AND fecha_fin <= DATEADD(DAY, 30, CAST(GETDATE() AS DATE))")["n"] \
        if cfg["notif_inventario"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} licencia(s) que vencen en los próximos 30 días",
                       "url": "/licencias?estado=Activa", "tipo": "alerta"})
    try:
        n = query_one(
            "SELECT COUNT(*) n FROM actividades WHERE completada = 0 "
            "AND proxima_accion IS NOT NULL AND proxima_accion <= CAST(GETDATE() AS DATE)")["n"] \
            if cfg.get("notif_actividades") == "1" else 0
    except Exception:
        n = 0
        # try/except por si la columna o el parametro no existieran en
        # una instalacion mas antigua: la notificacion falla en silencio
        # en vez de tumbar toda la barra superior.
    if n:
        avisos.append({"texto": f"{n} seguimiento(s) de clientes vencidos por realizar",
                       "url": "/actividades?pendientes=1", "tipo": "alerta"})
    n = query_one("SELECT COUNT(*) n FROM clientes WHERE estado = 'Prospecto'")["n"] \
        if cfg["notif_prospectos"] == "1" else 0
    if n:
        avisos.append({"texto": f"{n} cliente(s) prospecto por dar seguimiento",
                       "url": "/clientes?estado=Prospecto", "tipo": "info"})
    return jsonify({"total": len(avisos), "avisos": avisos})


"""
   Función: ayuda
   Objetivo: Mostrar la pantalla estática de ayuda del sistema.
   Parámetros: No recibe.
   Retorno: La plantilla ayuda.html.
"""
@bp.route("/ayuda")
def ayuda():
    return render_template("ayuda.html", active="")


"""
   Función: perfil
   Objetivo: Mostrar los datos del usuario que tiene la sesión abierta
             (GET) y permitirle editarlos (POST): su nombre, su email
             y, si lo desea, su contraseña, exigiéndole la contraseña
             actual para confirmar ese cambio.
   Parámetros: No recibe directamente (lee session["usuario_id"] y el
               formulario cuando es POST).
   Retorno: En GET, la plantilla perfil.html con los datos actuales.
            En POST, redirige a la misma pantalla con un mensaje.
"""
@bp.route("/perfil", methods=["GET", "POST"])
def perfil():
    """Perfil del usuario conectado: cambiar su nombre, email o contraseña."""
    from werkzeug.security import generate_password_hash, check_password_hash
    from ..db import execute  # noqa: F811
    from ..seguridad import registrar_auditoria
    from flask import flash, redirect, url_for

    uid = session["usuario_id"]      # Id del usuario conectado; siempre existe gracias a exigir_sesion()
    yo = query_one(
        "SELECT id, username, nombre, email, password_hash FROM usuarios WHERE id = ?", [uid])
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        email = (request.form.get("email") or "").strip()
        clave_actual = request.form.get("clave_actual") or ""
        clave_nueva = request.form.get("clave_nueva") or ""
        if not nombre or not email:
            flash("El nombre y el email son obligatorios.", "error")
            return redirect(url_for("main.perfil"))
        if clave_nueva:      # Solo se procesa el cambio de clave si escribieron una nueva
            if not check_password_hash(yo["password_hash"], clave_actual):
                flash("La contraseña actual no es correcta.", "error")
                return redirect(url_for("main.perfil"))
                # Se exige la clave ACTUAL antes de permitir el cambio:
                # evita que alguien que encuentre una sesion abierta
                # (por ejemplo en un equipo compartido) pueda cambiarla
                # sin conocerla.
            if len(clave_nueva) < 6:
                flash("La contraseña nueva debe tener al menos 6 caracteres.", "error")
                return redirect(url_for("main.perfil"))
            execute("UPDATE usuarios SET password_hash=? WHERE id=?",
                    [generate_password_hash(clave_nueva), uid])
        execute("UPDATE usuarios SET nombre=?, email=?, updated_at=SYSDATETIME() WHERE id=?",
                [nombre, email, uid])
        session["usuario_nombre"] = nombre
        # Se actualiza tambien el nombre guardado en la sesion, para que
        # el cambio se vea de inmediato en la barra superior sin tener
        # que cerrar y volver a abrir sesion.
        registrar_auditoria("usuarios", "UPDATE", uid,
                            datos_nuevos={"perfil": "actualizado",
                                          "cambio_password": bool(clave_nueva)})
        flash("Perfil actualizado.", "ok")
        return redirect(url_for("main.perfil"))
    return render_template("perfil.html", active="", yo=yo)


"""
   Función: configuracion
   Objetivo: Mostrar la pantalla de configuración del sistema (GET) y
             guardar los cambios (POST): los interruptores de
             notificación, el porcentaje de ITBIS, el tiempo de
             expiración de sesión, los datos del servidor de correo y
             las plantillas de los mensajes.
   Parámetros: No recibe directamente (lee el formulario si es POST).
   Retorno: En GET, la plantilla configuracion.html con los valores
            actuales. En POST, redirige a la misma pantalla.
"""
@bp.route("/configuracion", methods=["GET", "POST"])
def configuracion():
    """Configuración del sistema: notificaciones e impuesto. Solo administración."""
    from flask import flash, redirect, url_for
    from ..configuracion import leer_config, guardar_config, DEFECTOS
    from ..seguridad import registrar_auditoria

    if "usuarios.editar" not in session.get("permisos", []):
        flash("No tienes permiso para acceder a la configuración.", "error")
        return redirect(url_for("main.index"))
    # Esta pantalla NO usa el decorador @permiso_requerido: la
    # verificacion se hace a mano aqui dentro, con el mismo efecto.

    if request.method == "POST":
        nuevos = {
            "notif_activas": "1" if request.form.get("notif_activas") else "0",
            "notif_cotizaciones": "1" if request.form.get("notif_cotizaciones") else "0",
            "notif_facturas": "1" if request.form.get("notif_facturas") else "0",
            "notif_equipos": "1" if request.form.get("notif_equipos") else "0",
            "notif_prospectos": "1" if request.form.get("notif_prospectos") else "0",
            "notif_inventario": "1" if request.form.get("notif_inventario") else "0",
            # Cada casilla de verificacion llega con texto si esta marcada,
            # o no llega en absoluto si no lo esta: se traduce a "1" o "0"
            # como TEXTO, porque la columna configuracion.valor es VARCHAR.
            "itbis_pct": str(min(100, max(0, int(request.form.get("itbis_pct") or 18)))),
            # Doble acotacion: el impuesto no puede ser negativo ni mayor a 100.
            "notif_actividades": "1" if request.form.get("notif_actividades") else "0",
            "sesion_timeout_min": str(max(0, int(request.form.get("sesion_timeout_min") or 30))),
            "smtp_host": (request.form.get("smtp_host") or "").strip(),
            "smtp_puerto": (request.form.get("smtp_puerto") or "587").strip(),
            "smtp_usuario": (request.form.get("smtp_usuario") or "").strip(),
            "smtp_clave": request.form.get("smtp_clave") or "",
            "smtp_remitente": (request.form.get("smtp_remitente") or "").strip(),
            "plantilla_seguimiento": (request.form.get("plantilla_seguimiento") or "").strip(),
            "plantilla_renovacion": (request.form.get("plantilla_renovacion") or "").strip(),
        }
        try:
            guardar_config(nuevos)      # Usa la instruccion MERGE (ver configuracion.py)
            registrar_auditoria("configuracion", "UPDATE", None, datos_nuevos=nuevos)
            flash("Configuración guardada.", "ok")
        except Exception:
            flash("No se pudo guardar. Ejecuta primero sql/configuracion.sql en SSMS.", "error")
            # Si la tabla configuracion aun no existe en la base, el MERGE
            # falla; este mensaje le indica al usuario como solucionarlo.
        return redirect(url_for("main.configuracion"))

    return render_template("configuracion.html", active="config", cfg=leer_config())


# =====================================================================
#  EXPORTACIÓN DEL DASHBOARD (PDF / Excel)
# =====================================================================
"""
   Función: _datos_reporte
   Objetivo: Reunir en un solo lugar los datos que necesitan TANTO la
             exportación a Excel como la exportación a PDF, para no
             repetir las mismas consultas en las dos funciones.
   Parámetros: No recibe.
   Retorno: Tupla (kpis, fact_mes, top_cli): un diccionario con los
            cinco indicadores, la lista de facturación por mes de los
            últimos 6 meses, y la lista de los 10 clientes que más
            han facturado.
"""
def _datos_reporte():
    """Reúne los datos del dashboard para exportar."""
    kpis = {
        "Clientes activos": query_one("SELECT COUNT(*) n FROM clientes WHERE estado='Activo'")["n"],
        "Clientes totales": query_one("SELECT COUNT(*) n FROM clientes")["n"],
        "Cotizaciones pendientes": query_one(
            "SELECT COUNT(*) n FROM cotizaciones WHERE estado IN ('Borrador','Enviada')")["n"],
        "Equipos operativos": query_one(
            "SELECT COUNT(*) n FROM equipos WHERE estado='Operativo'")["n"],
        "Ingresos ultimos 30 dias (USD)": float(query_one(
            "SELECT COALESCE(SUM(total),0) s FROM facturas WHERE estado='Pagada' "
            "AND fecha_emision >= DATEADD(DAY,-30,CAST(GETDATE() AS DATE))")["s"]),
    }
    # Los cinco indicadores del reporte, en un diccionario cuyas claves
    # son directamente los textos que apareceran en el archivo exportado.
    fact_mes = query_all(
        "SELECT FORMAT(fecha_emision,'yyyy-MM') mes, SUM(total) total, COUNT(*) n "
        "FROM facturas WHERE estado='Pagada' AND fecha_emision >= DATEADD(MONTH,-6,GETDATE()) "
        "GROUP BY FORMAT(fecha_emision,'yyyy-MM') ORDER BY mes")
    top_cli = query_all(
        "SELECT TOP 10 e.nombre, SUM(f.total) total FROM facturas f "
        "JOIN clientes c ON c.id=f.cliente_id JOIN empresas e ON e.id=c.empresa_id "
        "WHERE f.estado='Pagada' GROUP BY e.nombre ORDER BY total DESC")
    return kpis, fact_mes, top_cli


"""
   Función: dashboard_excel
   Objetivo: Generar un archivo Excel (.xlsx) de tres hojas —
             indicadores, facturación mensual y top de clientes — con
             los mismos datos del dashboard, para descargarlo.
   Parámetros: No recibe.
   Retorno: Una respuesta HTTP con el archivo .xlsx como descarga.
"""
@bp.route("/dashboard/exportar/excel")
def dashboard_excel():
    from flask import Response
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    kpis, fact_mes, top_cli = _datos_reporte()      # Reutiliza las mismas consultas que el PDF
    wb = Workbook()                                  # Libro de Excel nuevo, construido en memoria
    azul = PatternFill("solid", fgColor="15639E")    # Color de fondo de los encabezados
    blanco = Font(color="FFFFFF", bold=True)         # Texto blanco y en negrita para los encabezados

    ws = wb.active; ws.title = "KPIs"                # Primera hoja: los indicadores
    ws.append(["Indicador", "Valor"])                 # Fila de encabezado
    for c in ws[1]: c.fill, c.font = azul, blanco     # Aplica el estilo a toda la fila 1
    for k, v in kpis.items(): ws.append([k, v])       # Una fila por cada indicador
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("Facturacion mensual")      # Segunda hoja
    ws2.append(["Mes", "Total (USD)", "Facturas"])
    for c in ws2[1]: c.fill, c.font = azul, blanco
    for f in fact_mes: ws2.append([f["mes"], float(f["total"]), f["n"]])
    ws2.column_dimensions["A"].width = 12; ws2.column_dimensions["B"].width = 16

    ws3 = wb.create_sheet("Top clientes")             # Tercera hoja
    ws3.append(["Cliente", "Facturado (USD)"])
    for c in ws3[1]: c.fill, c.font = azul, blanco
    for f in top_cli: ws3.append([f["nombre"], float(f["total"])])
    ws3.column_dimensions["A"].width = 34; ws3.column_dimensions["B"].width = 18

    buf = io.BytesIO(); wb.save(buf)                  # Guarda el libro en memoria, no en disco
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_dashboard.xlsx"})
    # El mimetype identifica el archivo como Excel moderno (.xlsx), y
    # Content-Disposition le indica al navegador que lo descargue en vez
    # de intentar mostrarlo en pantalla.


"""
   Función: dashboard_pdf
   Objetivo: Generar un archivo PDF con los indicadores, la
             facturación mensual y el top de clientes, dibujado a mano
             con la librería reportlab (posicionando cada texto por
             coordenadas exactas, no convirtiendo una página HTML).
   Parámetros: No recibe.
   Retorno: Una respuesta HTTP con el archivo .pdf como descarga.
"""
@bp.route("/dashboard/exportar/pdf")
def dashboard_pdf():
    from flask import Response
    import io, datetime as dt
    from reportlab.lib.pagesizes import letter    # Tamaño de pagina carta
    from reportlab.lib.units import cm             # Permite medir en centimetros en vez de puntos
    from reportlab.pdfgen import canvas            # El "lienzo" sobre el que se dibuja

    kpis, fact_mes, top_cli = _datos_reporte()
    buf = io.BytesIO(); c = canvas.Canvas(buf, pagesize=letter)   # PDF construido en memoria
    an, al = letter                                 # Ancho y alto de la pagina, en puntos
    # --- Encabezado azul con el titulo del reporte ---
    c.setFillColorRGB(0.08, 0.39, 0.62); c.rect(0, al-2.2*cm, an, 2.2*cm, fill=1, stroke=0)
    c.setFillColorRGB(1,1,1); c.setFont("Helvetica-Bold", 18)
    c.drawString(2*cm, al-1.5*cm, "CIBERSEG - Reporte ejecutivo")
    c.setFont("Helvetica", 10)
    c.drawRightString(an-2*cm, al-1.5*cm, dt.date.today().strftime("%d/%m/%Y"))
    c.setFillColorRGB(0.12,0.16,0.23)
    y = al-3.4*cm     # 'y' es la posicion vertical actual: se va bajando a mano, linea por linea
    # --- Bloque de indicadores ---
    c.setFont("Helvetica-Bold", 13); c.drawString(2*cm, y, "Indicadores"); y -= 0.7*cm
    c.setFont("Helvetica", 10)
    for k, v in kpis.items():
        val = f"${v:,.2f}" if "USD" in k else str(v)   # Solo los montos llevan formato de moneda
        c.drawString(2*cm, y, k); c.drawRightString(11*cm, y, val); y -= 0.55*cm
    y -= 0.5*cm
    # --- Bloque de facturacion mensual ---
    c.setFont("Helvetica-Bold", 13); c.drawString(2*cm, y, "Facturacion (6 meses)"); y -= 0.7*cm
    c.setFont("Helvetica", 10)
    for f in fact_mes:
        c.drawString(2*cm, y, f["mes"]); c.drawRightString(9*cm, y, f"${float(f['total']):,.2f}")
        c.drawRightString(11*cm, y, f"{f['n']} fact."); y -= 0.5*cm
    y -= 0.5*cm
    # --- Bloque del top de clientes ---
    c.setFont("Helvetica-Bold", 13); c.drawString(2*cm, y, "Top clientes"); y -= 0.7*cm
    c.setFont("Helvetica", 10)
    for f in top_cli:
        c.drawString(2*cm, y, f["nombre"][:45]); c.drawRightString(11*cm, y, f"${float(f['total']):,.2f}")
        y -= 0.5*cm
        if y < 2.5*cm: c.showPage(); y = al-2.5*cm; c.setFont("Helvetica", 10)
        # Si la posicion vertical se acerca al pie de pagina, se cierra
        # la pagina actual (showPage) y se abre una nueva, reiniciando 'y'.
    c.setFont("Helvetica", 8); c.setFillColorRGB(0.55,0.6,0.67)
    c.drawString(2*cm, 1.5*cm, "Generado por CIBERSEG")     # Pie de pagina
    c.save()                                                  # Cierra el PDF y lo deja listo en el buffer
    return Response(buf.getvalue(), mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_dashboard.pdf"})
